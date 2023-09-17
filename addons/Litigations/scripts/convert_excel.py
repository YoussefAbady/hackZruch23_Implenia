import openpyxl


file_path = "C:/Users/yyahiya/Downloads/input_Sample.xlsx"
new_file = "C:/Users/yyahiya/Desktop/test.txt"


dataframe = openpyxl.load_workbook(file_path)
f_n = open(new_file, "w",encoding='UTF-8')

dataframe1 = dataframe.active
start_row = 11
 
for row in range(start_row, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)








